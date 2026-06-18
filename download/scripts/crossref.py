#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Crossref Metadata Search — 搜索 + OA PDF 下载

功能：
  1. 在 search.crossref.org 搜索论文（表单提交）
  2. 文本解析提取标题/DOI/期刊/年份
  3. 筛选年份 + 管理学期刊优先级
  4. 通过 Crossref API + DOI 详情页获取 PDF 下载链接
  5. expect_download 浏览器下载 → 保存 PDF
  6. 失败自动记录 Excel

用法：
  python scripts/main.py crossref "关键词 | 年份范围 | 数量 | 下载目录"
  python scripts/main.py crossref "fintech prediction | 2025 2026 | 5 | D:/md"

参数格式（管道符分隔）：
  1st: 搜索关键词（必填）
  2nd: 年份范围 "startYear endYear"（可选，默认 2025 2026）
  3rd: 下载数量（可选，默认 5）
  4th: 下载目录（可选，默认 D:/md）
"""

import sys
import os
import re
import json
import time
import urllib.request
import urllib.parse

from utils import sp, safe_filename, parse_pipe_args, ensure_output_dir


# ── 默认配置 ──────────────────────────────────────────────────────────────
DEFAULT_COUNT = 5
DEFAULT_OUTPUT = "D:/md"


# ── 管理学期刊关键词（UTD24 + ABS 3*+ 英文 + 中文顶刊）────────────────────
GOOD_VENUE_KEYWORDS = [
    "management science", "operations research", "manufacturing & service operations",
    "production and operations management", "journal of operations management",
    "organization science", "academy of management", "administrative science quarterly",
    "strategic management journal", "journal of international business studies",
    "information systems research", "mis quarterly", "journal of management",
    "decision sciences", "decision support systems", "european journal of operational research",
    "omega", "computers & operations research", "expert systems with applications",
    "technological forecasting", "international journal of forecasting", "journal of forecasting",
    "futures", "annals of operations research",
    "ieee transactions on engineering management", "research policy",
    "journal of business research", "technovation",
    "international journal of production research", "international journal of production economics",
    "中国管理科学", "管理世界", "系统工程理论与实践", "管理科学学报",
    "管理工程学报", "管理评论", "管理学报", "科研管理",
    "科学学研究", "系统工程学报", "系统管理学报", "运筹与管理",
    "中国软科学", "预测", "管理科学",
]


def is_good_venue(venue_text):
    """判断是否为目标管理学期刊"""
    if not venue_text:
        return False
    vt = venue_text.lower()
    return any(v in vt for v in GOOD_VENUE_KEYWORDS)


def get_oa_from_crossref_api(doi):
    """通过 Crossref API 获取 OA/PDF 链接"""
    if not doi:
        return ""
    doi_clean = doi.replace("https://doi.org/", "").replace("http://dx.doi.org/", "").split("?")[0]
    url = f"https://api.crossref.org/works/{urllib.parse.quote(doi_clean)}"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (Academic)"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read().decode())
        msg = data.get("message", {})

        for link in msg.get("link", []):
            ct = link.get("content-type", "")
            u = link.get("URL", "")
            if ct == "application/pdf" and u:
                return u
            if u and ("pdf" in u.lower() or "/pdf/" in u):
                return u

        res = msg.get("resource", {})
        if res.get("primary", {}).get("URL"):
            return res["primary"]["URL"]
        return ""
    except Exception:
        return ""


def search_papers_from_crossref(page, keyword):
    """在 Crossref search 页搜索论文，返回提取的论文列表"""
    page.goto("https://search.crossref.org/", wait_until="domcontentloaded", timeout=30000)
    time.sleep(3)

    page.fill("#search-input", keyword)
    page.keyboard.press("Enter")

    try:
        page.wait_for_function('document.body.innerText.includes("results")', timeout=20000)
    except:
        pass
    time.sleep(5)

    text = page.evaluate("document.body.innerText")
    lines = [l.strip() for l in text.split("\n") if l.strip()]

    papers = []
    i = 0
    while i < len(lines):
        line = lines[i]
        if "doi.org" in line.lower() and "10." in line:
            doi = line.strip()
            title = venue = ""
            year = 0

            for back in range(1, 5):
                if i - back >= 0:
                    prev = lines[i - back]
                    m = re.match(r'([A-Z\s]+)\s+published\s+(\d{4})\s+in\s+(.+)', prev)
                    if m:
                        year = int(m.group(2))
                        venue = m.group(3).strip()
                    else:
                        m2 = re.search(r'published\s+.*?(\d{4})\s+in\s+(.+)', prev)
                        if m2:
                            year = int(m2.group(1))
                            venue = m2.group(2).strip()

            for back in range(1, 5):
                if i - back >= 0:
                    prev = lines[i - back]
                    if len(prev) > 15 and not prev.startswith("https") and \
                       "author" not in prev.lower() and "published" not in prev.lower():
                        title = prev
                        break

            papers.append({"title": title or doi, "doi": doi, "venue": venue, "year": year})
        i += 1

    # 去重
    seen = set()
    unique = []
    for p in papers:
        key = p["title"].lower()[:40]
        if key not in seen:
            seen.add(key)
            unique.append(p)

    return unique


def get_pdf_url_from_doi(page, doi_clean):
    """导航到 DOI 页面提取真实 PDF 下载链接"""
    try:
        doi_url = f"https://doi.org/{doi_clean}"
        page.goto(doi_url, wait_until="domcontentloaded", timeout=20000)
        time.sleep(3)

        pdf_info = page.evaluate("""
            () => {
                const url = document.location.href;
                let pdfLink = '';

                const meta = document.querySelector('meta[name="citation_pdf_url"]');
                if (meta) pdfLink = meta.getAttribute('content');

                document.querySelectorAll('a[href]').forEach(a => {
                    const h = a.getAttribute('href') || '';
                    const t = a.textContent.toLowerCase();
                    if ((t.includes('download pdf') || t.includes('download full text')) && !pdfLink) {
                        pdfLink = h.startsWith('http') ? h : new URL(h, url).href;
                    }
                    if ((h.includes('.pdf') || h.includes('/pdf/')) && /download|pdf/.test(t) && !pdfLink) {
                        pdfLink = h.startsWith('http') ? h : new URL(h, url).href;
                    }
                });

                document.querySelectorAll('a[href*="/pdf"]').forEach(a => {
                    if (!pdfLink) {
                        const h = a.getAttribute('href');
                        pdfLink = h.startsWith('http') ? h : new URL(h, url).href;
                    }
                });

                return { currentUrl: url, pdfLink };
            }
        """)

        if pdf_info.get("pdfLink"):
            pdf_url = pdf_info["pdfLink"]
            if not pdf_url.startswith("http"):
                pdf_url = "https:" + pdf_url
            return pdf_url

        # Fallback: Crossref API
        return get_oa_from_crossref_api(doi_clean)
    except Exception:
        return get_oa_from_crossref_api(doi_clean)


def download_pdf(page, pdf_url, fpath):
    """通过 expect_download 下载 PDF 到指定路径"""
    with page.expect_download(timeout=20000) as dl_info:
        page.evaluate(f'window.location.href = "{pdf_url}"')

    download = dl_info.value
    download.save_as(fpath)
    return os.path.getsize(fpath)


def save_failed_excel(failed, output_dir):
    """保存失败记录到 Excel"""
    xlsx_path = os.path.join(output_dir, "failed_crossref.xlsx")
    try:
        from openpyxl import Workbook, styles
        wb = Workbook()
        ws = wb.active
        ws.title = "失败记录"
        headers = ["文章名称", "期刊", "DOI", "链接", "失败原因"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h).font = styles.Font(bold=True)
        for ri, f_rec in enumerate(failed, 2):
            ws.cell(row=ri, column=1, value=f_rec.get("title", ""))
            ws.cell(row=ri, column=2, value=f_rec.get("venue", ""))
            ws.cell(row=ri, column=3, value=f_rec.get("doi", ""))
            ws.cell(row=ri, column=4, value=f_rec.get("link", ""))
            ws.cell(row=ri, column=5, value=f_rec.get("reason", ""))
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 30
        wb.save(xlsx_path)
        sp(f"\n失败记录已保存: {xlsx_path}")
    except Exception as e:
        sp(f"  Excel 保存失败: {e}")
        csv_path = os.path.join(output_dir, "failed_crossref.csv")
        with open(csv_path, "w", encoding="utf-8-sig") as f:
            f.write("文章名称,期刊,DOI,链接,失败原因\n")
            for f_rec in failed:
                f.write(f"{f_rec.get('title','')},{f_rec.get('venue','')},{f_rec.get('doi','')},{f_rec.get('link','')},{f_rec.get('reason','')}\n")
        sp(f"  CSV 已保存: {csv_path}")


def main(args_text=""):
    # ── 解析参数 ──────────────────────────────────────────────────────────
    defaults = {
        "keyword": "fintech prediction",
        "start_year": 2025,
        "end_year": 2026,
        "count": DEFAULT_COUNT,
        "output_dir": DEFAULT_OUTPUT,
    }
    params = parse_pipe_args(args_text, defaults)

    keyword = params.get("keyword", defaults["keyword"])
    start_year = params.get("start_year", defaults["start_year"])
    end_year = params.get("end_year", defaults["end_year"])
    count = params.get("count", defaults["count"])
    output_dir = params.get("output_dir", defaults["output_dir"])

    sp("=" * 60)
    sp("Crossref Metadata Search 搜索下载")
    sp("=" * 60)
    sp(f"  关键词: {keyword}")
    sp(f"  年份: {start_year}-{end_year}")
    sp(f"  目标数量: {count}")
    sp(f"  下载目录: {output_dir}")

    # ── 1. 连接 Chrome ──────────────────────────────────────────────────
    sp("\n[1/5] 连接 Chrome CDP...")
    from playwright.sync_api import sync_playwright

    playwright = sync_playwright().start()
    browser = playwright.chromium.connect_over_cdp("http://localhost:9222")
    ctx = browser.contexts[0]
    page = ctx.new_page()
    sp("  页面已创建")

    # ── 2. 搜索 + 提取 ──────────────────────────────────────────────────
    sp("\n[2/5] 搜索论文...")
    papers = search_papers_from_crossref(page, keyword)
    sp(f"  提取到 {len(papers)} 篇论文")
    for p in papers[:10]:
        sp(f"    [{p.get('year','')}] {p['title'][:60]}  {p.get('venue','')[:40]}")

    # ── 3. 筛选年份 + 期刊排序 ──────────────────────────────────────────
    sp("\n[3/5] 筛选条件...")
    filtered = [p for p in papers if p.get("year") and start_year <= p["year"] <= end_year]
    good = [p for p in filtered if is_good_venue(p.get("venue", ""))]
    other = [p for p in filtered if not is_good_venue(p.get("venue", ""))]
    ordered = good + other
    sp(f"  年份 {start_year}-{end_year} 筛选后: {len(filtered)} 篇")
    sp(f"  管理学期刊: {len(good)} 篇")

    # ── 4. 获取 OA PDF 链接 ─────────────────────────────────────────────
    sp("\n[4/5] 获取 OA 链接...")

    oa_papers = []
    for paper in ordered[:15]:
        doi = paper.get("doi", "")
        if not doi:
            continue
        doi_clean = doi.replace("https://doi.org/", "").replace("http://dx.doi.org/", "").split("?")[0]
        paper["doi_clean"] = doi_clean
        sp(f"  [{paper['title'][:45]}] ...")

        # 先访问 DOI 页找真实 PDF 链接
        pdf_url = get_pdf_url_from_doi(page, doi_clean)
        if pdf_url:
            paper["pdf_url"] = pdf_url
            oa_papers.append(paper)
            sp(f"    PDF: {pdf_url[:70]}...")
        else:
            sp(f"    无可用 PDF 链接")

    sp(f"  共获取到 {len(oa_papers)} 篇可下载论文")

    # ── 5. 下载 ─────────────────────────────────────────────────────────
    sp("\n[5/5] 下载 PDF...")
    ensure_output_dir(output_dir)

    downloaded = []
    failed = []
    processed = 0

    for paper in oa_papers:
        if processed >= count:
            break

        title = paper["title"]
        pdf_url = paper.get("pdf_url", "")
        sp(f"\n  --- {processed+1}/{count}: {title[:50]}...")

        if not pdf_url:
            failed.append({"title": title, "venue": paper.get("venue", ""),
                           "doi": paper.get("doi", ""), "link": "",
                           "reason": "未找到 PDF 链接"})
            continue

        if not pdf_url.startswith("http"):
            pdf_url = "https:" + pdf_url

        sp(f"  PDF: {pdf_url[:80]}...")
        fname = safe_filename(title, 60) + ".pdf"
        fpath = os.path.join(output_dir, fname)

        # 独立页面下载，避免页面状态干扰
        dl_page = ctx.new_page()
        try:
            size = download_pdf(dl_page, pdf_url, fpath)
            sp(f"  [OK] -> {fname} ({size} bytes)")
            downloaded.append({"title": title, "file": fname})
            processed += 1
        except Exception as e:
            sp(f"  [FAIL] {e}")
            failed.append({"title": title, "venue": paper.get("venue", ""),
                           "doi": paper.get("doi", ""), "link": pdf_url,
                           "reason": str(e)[:100]})
        finally:
            dl_page.close()
            time.sleep(1)

    # ── 结果汇总 ─────────────────────────────────────────────────────────
    sp("\n" + "=" * 60)
    sp("结果汇总")
    sp("=" * 60)
    sp(f"下载成功: {len(downloaded)} 篇")
    for d in downloaded:
        sp(f"  [OK] {d['title'][:60]}")

    sp(f"\n下载失败: {len(failed)} 篇")
    for f_rec in failed:
        sp(f"  [FAIL] {f_rec['title'][:60]}")
        sp(f"         原因: {f_rec.get('reason', '')[:60]}")

    if failed:
        save_failed_excel(failed, output_dir)

    playwright.stop()
    sp("\n完成!")


if __name__ == "__main__":
    args = " ".join(sys.argv[1:]) if len(sys.argv) > 1 else ""
    main(args)
