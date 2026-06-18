#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Semantic Scholar — 搜索 + OA筛选 + 批量下载

功能：
  1. 通过浏览器导航到 Semantic Scholar 搜索论文
  2. URL参数筛选年份 + 页面点击 OA 筛选
  3. 逐篇检查详情页 → 提取 PDF 链接
  4. 浏览器内 fetch → base64 → Python 解码保存 PDF
  5. 优先管理学期刊（UTD24 / ABS 3*+ / 中文顶刊）
  6. 失败自动记录 Excel

用法：
  python scripts/main.py semantic "关键词 | 年份范围 | 数量 | 下载目录"
  python scripts/main.py semantic "fintech prediction | 2025 2026 | 5 | D:/md"

参数格式（管道符分隔）：
  1st: 搜索关键词（必填）
  2nd: 年份范围 "startYear endYear"（可选，默认 2025 2026）
  3rd: 下载数量（可选，默认 5）
  4th: 下载目录（可选，默认 D:/md）
"""

import sys
import os
import re
import json
import time
import base64
import urllib.parse

from utils import sp, safe_filename, parse_pipe_args, ensure_output_dir


# ── 默认配置 ──────────────────────────────────────────────────────────────
DEFAULT_COUNT = 5
DEFAULT_OUTPUT = "D:/md"

# ── 管理学期刊关键词（UTD24 + ABS 3*+ 英文 + 中文顶刊）────────────────────
GOOD_VENUE_KEYWORDS = [
    "management science", "operations research", "manufacturing & service operations",
    "production and operations management", "journal of operations management",
    "organization science", "academy of management", "administrative science quarterly",
    "strategic management journal", "journal of international business studies",
    "information systems research", "mis quarterly", "journal of management",
    "decision sciences", "decision support systems", "european journal of operational research",
    "omega", "computers & operations research", "expert systems with applications",
    "technological forecasting", "international journal of forecasting", "journal of forecasting",
    "futures", "annals of operations research",
    "ieee transactions on engineering management", "research policy",
    "journal of business research", "technovation",
    "international journal of production research", "international journal of production economics",
    "中国管理科学", "管理世界", "系统工程理论与实践", "管理科学学报",
    "管理工程学报", "管理评论", "管理学报", "科研管理",
    "科学学研究", "系统工程学报", "系统管理学报", "运筹与管理",
    "中国软科学", "预测", "管理科学",
]


def is_good_venue(venue_text):
    """判断是否为目标管理学期刊"""
    if not venue_text:
        return False
    vt = venue_text.lower()
    return any(v in vt for v in GOOD_VENUE_KEYWORDS)


def wait_for_render(page, timeout=20000):
    """等待 Semantic Scholar React 页面渲染完成"""
    try:
        page.wait_for_function(
            "document.querySelectorAll('a[href*=\"/paper/\"]').length > 3",
            timeout=timeout
        )
        return True
    except:
        return False


def main(args_text=""):
    # ── 解析参数 ──────────────────────────────────────────────────────────
    defaults = {
        "keyword": "fintech prediction",
        "start_year": 2025,
        "end_year": 2026,
        "count": DEFAULT_COUNT,
        "output_dir": DEFAULT_OUTPUT,
    }
    params = parse_pipe_args(args_text, defaults)

    keyword = params.get("keyword", defaults["keyword"])
    start_year = params.get("start_year", defaults["start_year"])
    end_year = params.get("end_year", defaults["end_year"])
    count = params.get("count", defaults["count"])
    output_dir = params.get("output_dir", defaults["output_dir"])

    sp("=" * 60)
    sp("Semantic Scholar 搜索下载")
    sp("=" * 60)
    sp(f"  关键词: {keyword}")
    sp(f"  年份: {start_year}-{end_year}")
    sp(f"  目标数量: {count}")
    sp(f"  下载目录: {output_dir}")

    # ── 1. 连接 Chrome CDP ────────────────────────────────────────────────
    sp("\n[1/5] 连接 Chrome CDP...")
    from playwright.sync_api import sync_playwright

    playwright = sync_playwright().start()
    browser = playwright.chromium.connect_over_cdp("http://localhost:9222")
    ctx = browser.contexts[0]
    page = ctx.new_page()
    sp("  页面已创建")

    # ── 2. 搜索 ────────────────────────────────────────────────────────────
    sp("\n[2/5] 搜索论文...")
    query = urllib.parse.quote(keyword)
    url = f"https://www.semanticscholar.org/search?q={query}&sort=relevance"
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    time.sleep(5)

    if wait_for_render(page):
        sp("  搜索结果已渲染")
    else:
        sp("  等待渲染超时，继续...")
    sp(f"  页面标题: {page.title()}")

    # ── 3. 筛选条件 ──────────────────────────────────────────────────────
    sp("\n[3/5] 应用筛选条件...")

    # 3a. 年份筛选 — 直接改URL
    sp(f"  年份筛选 {start_year}-{end_year}...")
    years_param = "".join(f"&year%5B%5D={y}" for y in range(start_year, end_year + 1))
    filter_url = f"https://www.semanticscholar.org/search?q={query}&sort=relevance{years_param}"
    page.goto(filter_url, wait_until="domcontentloaded", timeout=60000)
    time.sleep(5)
    wait_for_render(page)

    # 3b. OA 筛选 — 尝试点击页面上的 Open Access 筛选器
    sp("  OA筛选...")
    try:
        oa_selectors = [
            "label:has-text('Open Access')",
            "span:has-text('Open Access')",
            "div:has-text('Open Access')",
            "text=Open Access",
        ]
        clicked = False
        for sel in oa_selectors:
            try:
                el = page.locator(sel).first
                if el.is_visible(timeout=3000):
                    el.click()
                    sp("    已点击 Open Access")
                    time.sleep(3)
                    clicked = True
                    break
            except:
                continue
        if not clicked:
            sp("    Open Access 筛选按钮不可见，继续全部预览")
    except Exception as e:
        sp(f"    OA 点击异常: {e}")

    time.sleep(3)

    # ── 4. 提取论文列表 ──────────────────────────────────────────────────
    sp("\n[4/5] 提取论文...")

    # 滚动加载更多
    for i in range(5):
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        time.sleep(1.5)

    # 从页面提取论文链接
    papers = page.evaluate("""
        () => {
            const links = document.querySelectorAll('a[href*="/paper/"]');
            const seen = new Set();
            const results = [];

            links.forEach(a => {
                const href = a.getAttribute('href');
                const title = (a.textContent || '').trim();

                if (!href || !title || title.length < 15 || seen.has(href)) return;
                seen.add(href);

                let parent = a.closest('[class*="result"], [class*="paper"], li, article') || a.parentElement;

                let venue = '';
                const venueEl = parent ? parent.querySelector('[class*="venue"], [class*="journal"]') : null;
                if (venueEl) venue = venueEl.textContent.trim();

                results.push({
                    title: title,
                    href: href.startsWith('http') ? href : 'https://www.semanticscholar.org' + href,
                    venue: venue,
                });
            });

            return results;
        }
    """)

    sp(f"  提取到 {len(papers)} 篇论文")
    for i, p in enumerate(papers[:10]):
        sp(f"    {i+1}. {p['title'][:60]}")
        if p.get('venue'):
            sp(f"       来源: {p['venue'][:40]}")

    # ── 逐篇检查 OA 状态 ──────────────────────────────────────────────────
    sp("\n  检查 OA 状态...")
    oa_papers = []
    for paper in papers[:30]:
        try:
            page.goto(paper['href'], wait_until="domcontentloaded", timeout=20000)
            time.sleep(2)
        except:
            continue

        info = page.evaluate("""
            () => {
                const body = (document.body.textContent || '').toLowerCase();
                const isOA = body.includes('open access') || body.includes('open-access');

                let venue = '';
                const metaEls = document.querySelectorAll('[class*="venue"], [class*="source"], [data-test-id="venue"]');
                metaEls.forEach(el => { if (el.textContent.trim()) venue = el.textContent.trim(); });

                const pdfUrls = [];
                document.querySelectorAll('a[href]').forEach(a => {
                    const h = a.getAttribute('href') || '';
                    const t = a.textContent.toLowerCase();
                    if (h.includes('.pdf') || t.includes('pdf') || t.includes('download fulltext')) {
                        pdfUrls.push(h.startsWith('http') ? h : 'https://www.semanticscholar.org' + h);
                    }
                });

                let doi = '';
                document.querySelectorAll('a[href*="doi.org"]').forEach(a => {
                    doi = a.getAttribute('href') || doi;
                });

                let oaPdfUrl = '';
                const meta = document.querySelector('meta[name="citation_pdf_url"]');
                if (meta) oaPdfUrl = meta.getAttribute('content');

                return { isOA, venue, pdfUrls, doi, oaPdfUrl };
            }
        """)

        paper['isOA'] = info.get('isOA', False)
        paper['venue'] = info.get('venue', '') or paper.get('venue', '')
        paper['pdfUrls'] = info.get('pdfUrls', [])
        paper['doi'] = info.get('doi', '')
        paper['oaPdfUrl'] = info.get('oaPdfUrl', '')

        sp(f"    [{paper['title'][:50]}] OA={paper['isOA']} PDFs={len(paper.get('pdfUrls',[]))}")

        if paper['isOA'] or paper.get('oaPdfUrl') or paper.get('pdfUrls'):
            oa_papers.append(paper)

    # 优先管理学期刊
    good = [p for p in oa_papers if is_good_venue(p.get('venue', ''))]
    other = [p for p in oa_papers if not is_good_venue(p.get('venue', ''))]
    ordered = good + other

    sp(f"\n  OA 论文总计: {len(oa_papers)} 篇")
    sp(f"  管理学期刊: {len(good)} 篇")

    # ── 5. 下载 PDF ──────────────────────────────────────────────────────
    sp("\n[5/5] 下载 PDF...")
    ensure_output_dir(output_dir)
    downloaded = []
    failed = []
    processed = 0

    for paper in ordered:
        if processed >= count:
            break

        title = paper['title']
        sp(f"\n  --- {processed+1}/{count}: {title[:50]}...")

        # 确定 PDF 链接
        pdf_url = paper.get('oaPdfUrl', '') or ''
        if not pdf_url:
            for pu in paper.get('pdfUrls', []):
                if 'pdf' in pu.lower() or pu.endswith('.pdf'):
                    pdf_url = pu
                    break
        if not pdf_url and paper.get('pdfUrls'):
            pdf_url = paper['pdfUrls'][0]

        if not pdf_url:
            sp("  [SKIP] 无 PDF 链接")
            failed.append({
                "title": title, "venue": paper.get('venue', ''),
                "doi": paper.get('doi', ''), "link": paper.get('href', ''),
                "reason": "文章页未找到 PDF 链接"
            })
            continue

        sp(f"  PDF URL: {pdf_url[:80]}...")

        try:
            pdf_b64 = page.evaluate(f"""
                async () => {{
                    try {{
                        const r = await fetch('{pdf_url}', {{
                            credentials: 'include',
                            headers: {{ 'Accept': 'application/pdf,*/*' }}
                        }});
                        if (!r.ok) return 'HTTP_ERROR:' + r.status;
                        const b = await r.blob();
                        const reader = new FileReader();
                        return await new Promise(res => {{
                            reader.onloadend = () => res(reader.result.split(',')[1]);
                            reader.readAsDataURL(b);
                        }});
                    }} catch(e) {{ return 'FETCH_ERROR:' + e.message; }}
                }}
            """)

            if pdf_b64 and not pdf_b64.startswith("ERROR") and not pdf_b64.startswith("HTTP_"):
                fname = safe_filename(title, 60) + ".pdf"
                fpath = os.path.join(output_dir, fname)
                with open(fpath, "wb") as f:
                    f.write(base64.b64decode(pdf_b64))
                sp(f"  [OK] -> {fname}")
                downloaded.append({"title": title, "file": fname})
                processed += 1
            else:
                sp(f"  [FAIL] {str(pdf_b64)[:80]}")
                failed.append({
                    "title": title, "venue": paper.get('venue', ''),
                    "doi": paper.get('doi', ''), "link": paper.get('href', ''),
                    "reason": str(pdf_b64)[:100]
                })
        except Exception as e:
            sp(f"  [FAIL] {e}")
            failed.append({
                "title": title, "venue": paper.get('venue', ''),
                "doi": paper.get('doi', ''), "link": paper.get('href', ''),
                "reason": str(e)[:100]
            })

    # ── 结果汇总 ──────────────────────────────────────────────────────────
    sp("\n" + "=" * 60)
    sp("结果汇总")
    sp("=" * 60)
    sp(f"下载成功: {len(downloaded)} 篇")
    for d in downloaded:
        sp(f"  [OK] {d['title'][:60]}")

    sp(f"\n下载失败: {len(failed)} 篇")
    for f_rec in failed:
        sp(f"  [FAIL] {f_rec['title'][:60]}")
        sp(f"         原因: {f_rec.get('reason', '')[:60]}")

    # 保存失败记录 Excel
    if failed:
        xlsx_path = os.path.join(output_dir, "failed_semanticscholar.xlsx")
        try:
            from openpyxl import Workbook, styles
            wb = Workbook()
            ws = wb.active
            ws.title = "失败记录"
            headers = ["文章名称", "期刊", "DOI", "链接", "失败原因"]
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = styles.Font(bold=True)
            for ri, f_rec in enumerate(failed, 2):
                ws.cell(row=ri, column=1, value=f_rec.get("title", ""))
                ws.cell(row=ri, column=2, value=f_rec.get("venue", ""))
                ws.cell(row=ri, column=3, value=f_rec.get("doi", ""))
                ws.cell(row=ri, column=4, value=f_rec.get("link", ""))
                ws.cell(row=ri, column=5, value=f_rec.get("reason", ""))
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 30
            wb.save(xlsx_path)
            sp(f"\n失败记录已保存: {xlsx_path}")
        except Exception as e:
            sp(f"  Excel 保存失败，保存 CSV: {e}")
            csv_path = os.path.join(output_dir, "failed_semanticscholar.csv")
            with open(csv_path, "w", encoding="utf-8-sig") as f:
                f.write("文章名称,期刊,DOI,链接,失败原因\n")
                for f_rec in failed:
                    f.write(f"{f_rec.get('title','')},{f_rec.get('venue','')},{f_rec.get('doi','')},{f_rec.get('link','')},{f_rec.get('reason','')}\n")
            sp(f"   CSV 已保存: {csv_path}")

    playwright.stop()
    sp("\n完成!")


if __name__ == "__main__":
    args = " ".join(sys.argv[1:]) if len(sys.argv) > 1 else ""
    main(args)
